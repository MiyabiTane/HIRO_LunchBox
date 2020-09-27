#!/usr/bin/env python
# -*- coding: utf-8 -*-
import rospy
import actionlib
from speech_recognition_msgs.msg import SpeechRecognitionCandidates
from sound_play.msg import SoundRequest
from sound_play.libsoundplay import SoundClient
import random
import openpyxl

class RememberInfo:
    # 今はあらかじめ書いてあるおかずだけ
    # like : 1, dislike : 0
    # U : Unknown, N : None(重複)
    def __init__(self, wb_name):
        self.wb_name = wb_name
        self.like_list = []
        self.dislike_list = []
        self.want_to_eat = []
        self.wb = openpyxl.load_workbook(wb_name)
        self.sheet = self.wb['Sheet1']
        self.values = list(self.sheet.values)
        self.foods = [str(f) for f in self.values[0][1:]]
        self.N = len(self.foods)
        #print(self.values)
        #print(self.foods)

    def get_past_info(self):
        for i in range(1, self.N+1):
            for j in range(i, self.N+1):
                if self.values[i][j] == 1:
                    self.like_list.append([str(self.foods[i-1]), str(self.foods[j-1])])
                elif self.values[i][j] == 0:
                    self.dislike_list.append([str(self.foods[i-1]), str(self.foods[j-1])])
        
        for i,tf in enumerate(self.values[self.N+1][1:]):
            if str(tf) == 'T':
                self.want_to_eat.append(self.foods[i])
        return self.like_list, self.dislike_list, self.want_to_eat

    def update_info(self, new_like_list, new_dislike_list, new_want_to_eat):
        for i in range(len(new_like_list)):
            food1 = new_like_list[i][0]
            food2 = new_like_list[i][1]
            index1 = self.foods.index(food1) + 2
            index2 = self.foods.index(food2) + 2
            self.sheet.cell(row=min(index1,index2), column=max(index1,index2), value = 1)
       
        for i in range(len(new_dislike_list)):
            food1 = new_dislike_list[i][0]
            food2 = new_dislike_list[i][1]
            index1 = self.foods.index(food1) + 2
            index2 = self.foods.index(food2) + 2
            self.sheet.cell(row = min(index1, index2), column = max(index1, index2), value = 0)

        for i,food in enumerate(new_want_to_eat):
            index = self.foods.index(food)+2
            self.sheet.cell(row = self.N+2, column = index, value = 'T')
        self.wb.save(self.wb_name)

class TalkWith:
    
    def __init__(self):
        self.cb_flag = False
        self.voice = 'voice_kal_diphone'
        self.volume = 1.0
        self.str_candidate = ''
        self.like_list = []
        self.dislike_list = []
        self.want_to_eat = []
        #rospy.init_node('hiro_talk')
        self.soundhandle = SoundClient()
        rospy.sleep(1.0)

    def talk_cb(self, msg):  
        #print("talk_cb")
        if msg.transcript and not self.str_candidate:
            for char in msg.transcript:
                self.str_candidate += char
            print("got words {}".format(self.str_candidate))
            self.cb_flag = False

    def convertEtoJ(self, food):
        if food == "rolled_egg":
            return "卵焼き"
        elif food == "octopus_wiener":
            return "ウィンナー"
        elif food == "tomato":
            return "トマト"
        elif food == "broccoli":
            return "ブロッコリー"
        elif food == "flower_carrot":
            return "人参"
        elif food == "fried_chicken":
            return "唐揚げ"
        else:
            print("error")
            return

    def convertJtoE(self, food):
        if food == "卵焼き":
            return "rolled_egg"
        elif food == "ウィンナー":
            return "octopus_wiener"
        elif food == "トマト":
            return "tomato"
        elif food == "ブロッコリー":
            return "broccoli"
        elif food == "人参":
            return "flower_carrot"
        elif food == "唐揚げ":
            return "fried_chicken"
        else:
            print("error")
            return

    def main_before_stuff(self, name_list):
        """
        認識したおかずの情報を元に
        隣あって欲しい、欲しくないおかずの組み合わせ
        絶対にいれて欲しいおかず
        を記録
        """
        N = len(set(name_list))
        food_for_judge = []
        for _ in range(N):
            cand = random.sample(name_list, 2)
            while cand in food_for_judge or cand[::-1] in food_for_judge:
                cand = random.sample(name_list, 2)
            food_for_judge.append(cand)
        print("food_for_judge : ",food_for_judge)

        for food in food_for_judge:
            f1 = food[0]; f2 = food[1]
            #f1_ = self.convertEtoJ(f1); f2_ = self.convertEtoJ(f2)
            #s = f1_ + "と" + f2_ + "は隣で良いですか？"
            s = "Is it good that " + f1.replace('_',' ') + " and " + f2.replace('_',' ') + " are next to each other?"
            rospy.loginfo('Saying: %s' % s)
            self.soundhandle.say(s, self.voice, self.volume)
            
            while True:
                self.str_candidate = ""
                self.cb_flag = True
                while self.cb_flag:
                    self.sub = rospy.Subscriber('/speech_to_text', SpeechRecognitionCandidates, self.talk_cb)
                    rospy.sleep(0.5)
                if "はい" in self.str_candidate:
                    self.like_list.append([f1, f2])
                    break
                elif "いいえ" in self.str_candidate:
                    self.dislike_list.append([f1, f2])
                    break
                #s = "すみません、もう一度お願いします。"
                s = "Could you repeat that please?"
                rospy.loginfo('Saying: %s' % s)
                self.soundhandle.say(s, self.voice, self.volume)
        print(self.like_list, self.dislike_list)

        #s = "絶対に入れてほしいおかずをひとつずつ教えてください"
        s = "Please tell me one side dish you want to eat at a time."
        rospy.loginfo('Saying: %s' % s)
        self.soundhandle.say(s, self.voice, self.volume)
        while True:
            self.cb_flag = True
            self.str_candidate = ""
            while self.cb_flag:
                #print(self.cb_flag)
                self.sub = rospy.Subscriber('/speech_to_text', SpeechRecognitionCandidates, self.talk_cb)
                rospy.sleep(0.5)
            if  self.convertJtoE(self.str_candidate) in name_list:
                self.want_to_eat.append(self.convertJtoE(self.str_candidate))
                #s = self.str_candidate + "ですね、分かりました。他にはありますか？"
                s = "You want to eat " + self.convertJtoE(self.str_candidate).replace('_',' ') + " right?"
                rospy.loginfo('Saying: %s' % s)
                self.soundhandle.say(s, self.voice, self.volume)
                self.cb_flag = True
                self.str_candidate = ""
                while self.cb_flag:
                    self.sub = rospy.Subscriber('/speech_to_text', SpeechRecognitionCandidates, self.talk_cb)
                    rospy.sleep(0.5)
                if "はい" in self.str_candidate:
                    s = "OK. Is there anything else?"
                    rospy.loginfo('Saying: %s' % s)
                    self.soundhandle.say(s, self.voice, self.volume)
                elif "いいえ" in self.str_candidate:
                    s = "Could you repeat the food's name please?"
                    rospy.loginfo('Saying: %s' % s)
                    self.soundhandle.say(s, self.voice, self.volume)
                else:
                    s = "Could you repeat that please?"
                    rospy.loginfo('Saying: %s' % s)
                    self.soundhandle.say(s, self.voice, self.volume)
            elif "ない" in self.str_candidate or "ありません" in self.str_candidate or "いいえ"in self.str_candidate:
                #s = "了解しました。"
                s = "OK, I understand. Thank you."
                rospy.loginfo('Saying: %s' % s)
                self.soundhandle.say(s, self.voice, self.volume)
                self.sub.unregister()
                break
            else:
                #s = "すみません、それを用意することはできません、もう一度お願いします。"
                s = "I am sorry, I cannot prepare it. Is there anything else?"
                rospy.loginfo('Saying: %s' % s)
                self.soundhandle.say(s, self.voice, self.volume)
        
        # 過去の記憶を使う
        # To Do 重複質問をなくす
        rem = RememberInfo("owner_info.xlsx")
        rem.update_info(self.like_list, self.dislike_list, self.want_to_eat)
        rem = RememberInfo("owner_info.xlsx")
        new_like, new_dislike, new_want = rem.get_past_info()
        print(new_like)
        print(new_dislike)
        print(new_want)
        return new_like, new_dislike, new_want

"""
#talk
name_list = ["rolled_egg", "fried_chicken", "broccoli", "tomato", "octopus_wiener", "fried_chicken", "rolled_egg", "broccoli", "octopus_wiener", "tomato"]
hiro_talk = TalkWith()
hiro_talk.main_before_stuff(name_list)
"""
"""
#xlsx
new_like = [["rolled_egg", "rolled_egg"], ["tomato", "fried_chicken"]]
new_dislike = [["octopus_wiener", "fried_chicken"], ["rolled_egg","fried_chicken"]]
new_want = ["tomato", "rolled_egg"]
rem = RememberInfo("test.xlsx")
list1, list2, list3 = rem.get_past_info()
print("like : ",list1)
print("dislike : ",list2)
print("want : ",list3)
rem.update_info(new_like, new_dislike, new_want)
"""